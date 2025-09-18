from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os
import qrcode

app = Flask(__name__)

EXCEL_FILE = "registrations.xlsx"
QR_FOLDER = "static/qr_codes"

# Ensure folders exist
os.makedirs(QR_FOLDER, exist_ok=True)

# Initialize Excel if it doesn’t exist
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"
    ws.append(["Name", "Gender", "Phone Number", "QR Code URL"])
    wb.save(EXCEL_FILE)

# Route to generate QR for registration form
@app.route("/generate_qr")
def generate_qr():
    # Force PythonAnywhere domain
    domain = "https://seegesys25.pythonanywhere.com/"
    form_url = f"{domain}{url_for('form')}"

    qr_file_name = "registration_form.png"
    #qr_path = os.path.join(QR_FOLDER, qr_file_name)
    qr_path = os.path.join(app.root_path, "static", "qr_codes", qr_file_name)

    # Ensure folder exists
    os.makedirs(QR_FOLDER, exist_ok=True)

    # Generate QR code
    qr = qrcode.QRCode(version=1, box_size=8, border=2)
    qr.add_data(form_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(qr_path)

    # Public link to QR image
    qr_url = url_for("static", filename=f"qr_codes/{qr_file_name}", _external=True)

    return f"""
        <h2>Scan this QR to open the Registration Form</h2>
        <img src="{qr_url}" alt="QR Code" width="250">
        <p><a href="{form_url}" target="_blank">Open Form</a></p>
    """

# Registration form page
@app.route("/form")
def form():
    return render_template("form.html")

# Form submission
@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name")
    gender = request.form.get("gender")
    phone = request.form.get("phone")

    # Optional: generate a unique QR for this user
    reg_id = f"{name.replace(' ','')}{phone[-4:]}"
    qr_file_name = f"{reg_id}.png"
    qr_path = os.path.join(QR_FOLDER, qr_file_name)
    qr_data = f"Registration ID: {reg_id}\nName: {name}\nPhone: {phone}"
    qr = qrcode.QRCode(version=1, box_size=8, border=2)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill="black", back_color="white")
    img.save(qr_path)

    # Save to Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    qr_url = url_for('static', filename=f"qr_codes/{qr_file_name}", _external=True)
    ws.append([name, gender, phone, qr_url])
    wb.save(EXCEL_FILE)

    return f"""
        <h2>✅ Registration Successful!</h2>
        <p>Your QR code:</p>
        <img src="{qr_url}" alt="User QR" width="220">
        <p><a href='/form'>Register Another</a></p>
        <p><a href='{qr_url}' target='_blank'>Open QR Publicly</a></p>
    """

@app.route("/registrations")
def registrations():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))  # all rows
    headers = rows[0]
    data = rows[1:]

    # Build HTML table
    table_html = "<table border='1' cellpadding='6' style='border-collapse:collapse;'>"
    table_html += "<tr>" + "".join([f"<th>{h}</th>" for h in headers]) + "</tr>"

    for row in data:
        table_html += "<tr>" + "".join([f"<td>{cell}</td>" for cell in row]) + "</tr>"

    table_html += "</table>"

    return f"""
        <h2>📋 Registrations</h2>
        {table_html}
        <p><a href='/form'>⬅ Back to Form</a></p>
    """


if __name__ == "__main__":
    app.run(debug=True)
